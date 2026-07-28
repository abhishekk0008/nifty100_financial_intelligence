from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

class PeerComparisonReport:

    def __init__(
        self,
        ranking_df,
        companies_df,
    ):
        self.ranking_df = ranking_df.copy()
        self.companies_df = companies_df.copy()
        self.workbook = Workbook()

    def generate(self):

        # -----------------------------------------
        # Latest available record for EACH company
        # -----------------------------------------
        df = self.ranking_df.copy()

        df["year_num"] = (
            df["year"]
            .astype(str)
            .str.extract(r"(\d{4})")[0]
            .astype(int)
        )

        df = (
            df.sort_values(
                [
                    "company_id",
                    "metric",
                    "year_num",
                ]
            )
            .groupby(
                [
                    "company_id",
                    "metric",
                ],
                group_keys=False,
            )
            .tail(1)
            .drop(columns="year_num")
            .copy()
        )

        print("Latest record selected for each company")

        # -----------------------------------------
        # Merge company names
        # -----------------------------------------
        companies = (
            self.companies_df[
                [
                    "id",
                    "company_name",
                ]
            ]
            .rename(
                columns={
                    "id": "company_id",
                }
            )
        )

        df = df.merge(
            companies,
            on="company_id",
            how="left",
        )

        print("\nLong format")
        print(df.head())
        print(f"Rows: {len(df)}")
        print(f"Companies: {df['company_id'].nunique()}")
        print(f"Metrics: {df['metric'].nunique()}")

        # -----------------------------------------
        # Convert Long -> Wide
        # -----------------------------------------

        wide = (
        df.pivot(
        index=[
            "company_id",
            "company_name",
            "peer_group_name",
            "year",
        ],
        columns="metric",
        values=[
            "value",
            "percentile_rank",
        ],
    )
)

        # Flatten MultiIndex columns
        wide.columns = [
            f"{metric}_{kind}"
            for kind, metric in wide.columns
        ]

        wide = (
            wide.reset_index()
            .sort_values("company_id")
            .reset_index(drop=True)
        )

        print("\nWide format")
        print(wide.head())
        print(f"Wide shape: {wide.shape}")
        print(f"Companies: {wide['company_id'].nunique()}")
        print(f"Columns: {len(wide.columns)}")



        # -----------------------------------------
        # Create Workbook
        # -----------------------------------------

        # Remove default sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # Create output folder
        output_dir = Path("output")
        output_dir.mkdir(parents=True, exist_ok=True)

        # Get peer groups
        peer_groups = (
            wide["peer_group_name"]
            .fillna("No peer group assigned")
            .sort_values()
            .unique()
        )

        print(f"\nCreating {len(peer_groups)} worksheets...")

        # One worksheet per peer group
        for peer in peer_groups:

            sheet_name = str(peer)[:31]

            ws = self.workbook.create_sheet(
                title=sheet_name
            )

            peer_df = (
                wide[
                    wide["peer_group_name"]
                    .fillna("No peer group assigned")
                    == peer
                ]
                .sort_values("company_name")
                .reset_index(drop=True)
            )

            # Write dataframe
            for row in dataframe_to_rows(
                peer_df,
                index=False,
                header=True,
            ):
                ws.append(row)

            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Auto column width
            for column_cells in ws.columns:

                max_length = 0

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:

                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value)),
                        )

                ws.column_dimensions[
                    column_letter
                ].width = min(max_length + 2, 35)

        # Save workbook
        output_file = (
            output_dir
            / "peer_comparison.xlsx"
        )

        self.workbook.save(output_file)

        print(
            f"\nWorkbook saved to: {output_file}"
        )

        return wide